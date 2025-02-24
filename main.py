from datetime import datetime, date, timedelta
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from kivy.app import App
from kivy.uix.screenmanager import Screen, ScreenManager
from kivy.uix.label import Label
from kivy.uix.spinner import Spinner
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy_garden.matplotlib import FigureCanvasKivyAgg
from kivy.core.window import Window

def mostrar_popup(titulo, texto):
        msg=Popup(
            title = titulo,
            content = Label(text=texto),
            size_hint=(None, None),
            size = (300, 200),
            padding=(10, 10, 10, 10)
        )

        msg.open()

class Gerenciador(ScreenManager):
    pass

class Menu(Screen):
    pass

class Cadastro(Screen):
    pass

class CadastroTurmas(Screen):
    def on_pre_enter(self):
        self.ids.box.clear_widgets()
        try:
            arq_slvTurmas = openpyxl.load_workbook('cadastro_turmas.xlsx')
            celulas = arq_slvTurmas.active
            turmas_cadastradas = [celula.value for celula in celulas['A'][1:]]
            qtde_alunos = [celula.value for celula in celulas['B'][1:]]

        except FileNotFoundError:
            turmas_cadastradas = []
            qtde_alunos = []

        for turma, alunos in zip(turmas_cadastradas, qtde_alunos):
            self.ids.box.add_widget(Label(text=f'TURMA: {turma} - QUANTIDADE DE ALUNOS: {alunos}',
                                        size_hint_y=None,
                                        height=40))

    def salvar_dadosTurma(self, nome_turma, qtde_alunos):
        try:
            tab_slvDataTurma = openpyxl.load_workbook('cadastro_turmas.xlsx')
            celula = tab_slvDataTurma.active
        except FileNotFoundError:
            tab_slvDataTurma = openpyxl.Workbook()
            celula = tab_slvDataTurma.active
            celula.append(['Nome da Turma', 'Quantidade de Alunos'])

        if 0 < len(nome_turma) <= 10 and qtde_alunos.isdigit() and 1 <= int(qtde_alunos) <= 99:
            celula.append([nome_turma.upper(), int(qtde_alunos)])
            self.ids.box.add_widget(Label(text=f'TURMA: {nome_turma.upper()} - QUANTIDADE DE ALUNOS: {qtde_alunos}',
                                      size_hint_y=None,
                                      height=40))
            mostrar_popup('Sucesso!', "A Turma solicitada foi adicionada.")

        elif ValueError:
            if qtde_alunos == '' or nome_turma == '':
                #POPUP de campo vazio
                mostrar_popup('Campo Vazio!', 'Digite um valor válido. \nTurma deve ter no máximo 10 \ncaracteres; \nQuantidade de Pessoas deve \nser de 1 a 99.')
            else:
                mostrar_popup('Valor Inválido!', 'Digite um valor válido. \nTurma deve ter no máximo 10 \ncaracteres; \nQuantidade de Pessoas deve \nser de 1 a 99.')

        tab_slvDataTurma.save('cadastro_turmas.xlsx')
        self.ids.nomeTurma.text = ''
        self.ids.qtdeAlunos.text = ''

    def remover_turma(self, nome_turma):
        try:
            tab_slvDataTurma = openpyxl.load_workbook('cadastro_turmas.xlsx')
            celula = tab_slvDataTurma.active
        except FileNotFoundError:
            return mostrar_popup('Erro!', 'Não há turmas cadastradas. \nCadastre uma turma primeiro!')
        
        for row in celula.iter_rows(min_row=2, max_col=1):
            if row[0].value == nome_turma.upper():
                celula.delete_rows(row[0].row, 1)
                mostrar_popup('Sucesso!', 'A Turma solicitada foi removida.')
                break
        else:
            mostrar_popup('Campo Vazio ou Valor Inválido!', 'Essa turma não existe ou você \ndeixou o campo vazio. \nConfira as turmas cadastradas \nacima antes de remover.')

        tab_slvDataTurma.save('cadastro_turmas.xlsx')
        self.ids.nomeTurma.text = ''

        self.on_pre_enter()

class CadastroMonitor(Screen):
    def on_pre_enter(self): 
        self.ids.box.clear_widgets()
        try:
            arq_slvMonitor = openpyxl.load_workbook('cadastro_monitores.xlsx')
            celulas = arq_slvMonitor.active
            monitores_cadastrados = [celula.value for celula in celulas['A'][1:]]
        except FileNotFoundError:
            monitores_cadastrados = []

        for monitor in monitores_cadastrados:
            self.ids.box.add_widget(Label(text=monitor, size_hint_y=None, height=40))

    def salvar_monitor(self, nome_aluno):
        try:
            arq_slvMonitor = openpyxl.load_workbook('cadastro_monitores.xlsx')
            celula = arq_slvMonitor.active
        except FileNotFoundError:
            arq_slvMonitor = openpyxl.Workbook()
            celula = arq_slvMonitor.active
            celula.append(['Nome'])
        
        if len(nome_aluno)>= 3 and len(nome_aluno) <= 80:
            celula.append([nome_aluno.upper()])
            self.ids.box.add_widget(Label(text=nome_aluno.upper(), size_hint_y=None, height=40))
        else:
            mostrar_popup('Campo Vazio ou Valor Inválido!', 'Você digitou um nome inválido \nou deixou o campo vazio. \nDigite novamente!')

        arq_slvMonitor.save('cadastro_monitores.xlsx')
        self.ids.nomeAluno.text = ''
        
        self.on_pre_enter()

    def remover_monitor(self, nome_aluno):
        try:
            arq_slvMonitor = openpyxl.load_workbook('cadastro_monitores.xlsx')
            celula = arq_slvMonitor.active
        except FileNotFoundError:
            return mostrar_popup('Erro!', 'Não há monitores cadastrados. \nCadastre um monitor primeiro!')

        for row in celula.iter_rows(min_row=2, max_col=1):
            if row[0].value == nome_aluno.upper():
                celula.delete_rows(row[0].row, 1)
                mostrar_popup('Sucesso!', 'O Monitor solicitado foi removido.')
                break
        else:
            mostrar_popup('Erro de Digitação!', 'Esse nome não existe.\nConfira os monitores cadastrados \nacima antes de remover.')

        arq_slvMonitor.save('cadastro_monitores.xlsx')
        self.ids.nomeAluno.text = ''

        self.on_pre_enter()

class CadastroOrdem(Screen):
    def on_pre_enter(self):
        self.ids.grid.clear_widgets()
        try:
            arq_slvTurma = openpyxl.load_workbook('cadastro_turmas.xlsx')
            celulas = arq_slvTurma.active
            turmas_cadastradas = [celula.value for celula in celulas['A'][1:]]
                        
        except FileNotFoundError:
            mostrar_popup('Arquivo não encontrado!', 'Arquivo de Cadastro de Turmas não \nencontrado. Cadastre os dados \nfaltantes nas suas sessões primeiro.')
            return

        num_turmas = len(turmas_cadastradas)
        dias_semana = ['Turma','Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta'] 

        for dia in dias_semana:
            self.ids.grid.add_widget(Label(text=dia, size_hint=(0.2, None), height= 40))

        for turma in turmas_cadastradas:
            self.ids.grid.add_widget(Label(text=turma, size_hint=(0.2, None), height=40))
            for dia in dias_semana[1:]:
                    self.ids.grid.add_widget(Spinner(
                        text='Ordem',
                        values=[str(i) for i in range(1, num_turmas + 1)],
                        size_hint=(0.2, None),
                        size=(80, 40),
                        pos_hint={'center_x': 0.5}))

    def salvar_Ordem(self):
        try:
            arq_slvTurma = openpyxl.load_workbook('cadastro_turmas.xlsx')
            celula_turma = arq_slvTurma.active

            if celula_turma.max_row <= 1:
                mostrar_popup('Erro ao Salvar!', 'Não é possível salvar a Ordem. \nArquivo de Cadastro de Turmas está \nvazio. Cadastre as turmas primeiro.')
                return
            
        except FileNotFoundError:
            return mostrar_popup('Erro ao Salvar!', 'Não é possível salvar a Ordem. \nArquivo de Cadastro de Turmas está \nvazio. Cadastre as turmas primeiro.')
        
        try:
            arq_slvDataOrdem = openpyxl.load_workbook('cadastro_OrdemTurmas.xlsx')
            celula = arq_slvDataOrdem.active

        except FileNotFoundError:
            arq_slvDataOrdem = openpyxl.Workbook()
            celula = arq_slvDataOrdem.active
            celula.append(['Turma', 'Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta'])

        while celula.max_row > 1:
            celula.delete_rows(2, 1)
        
        divisao_grids = self.ids.grid.children[::-1]
        num_turmas = len(divisao_grids) // 6

        posicoes_usadas = {dia: set() for dia in ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta']}

        for index in range(num_turmas):
            turma_index = index * 6
            turma = divisao_grids[turma_index].text
            escolhas = []
            
            if not turma == 'Turma':
                for spn in range(1, 6):
                    spinner = divisao_grids[turma_index + spn]
                    if spinner.text.isdigit():
                        escolha = (int(spinner.text))
                        dia_semana = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta'][spn - 1]
                        if escolha in posicoes_usadas[dia_semana]:
                            mostrar_popup('Erro!', f'Duas turmas ou mais não podem estar \nna mesma posição na {dia_semana}.')
                            return
                        posicoes_usadas[dia_semana].add(escolha)
                        escolhas.append(escolha)
                    else:
                        mostrar_popup('Erro!', 'Todos os campos de ordem devem \nestar com números antes de salvar.')
                        return

                celula.append([turma] + escolhas)

        arq_slvDataOrdem.save('cadastro_OrdemTurmas.xlsx')
        for widget in divisao_grids:
            if isinstance(widget, Spinner):
                widget.text = 'Ordem'

class RegistroDia(Screen):
    def on_pre_enter(self):
        self.atualizar_monitores()
        self.ordenar_turmas_dia()
        
        self.data_dia = datetime.now()
        self.ids.datetimer.text = f"{self.data_dia.strftime('%d')}/{self.data_dia.strftime('%m')}/{self.data_dia.strftime('%Y')}"
        num = date.today().weekday()
        dias_semana = ('Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo')
        self.ids.diaSemana.text = dias_semana[num]
    
    def atualizar_monitores(self):
        try:
            tab_Monitor = openpyxl.load_workbook('cadastro_monitores.xlsx')
            celulas = tab_Monitor.active

            monitores_cadastrados = []
            for celula in celulas['A'][1:]:
                nome = celula.value.split(" ")
                monitores_cadastrados.append(nome[0]+' '+nome[-1])
            self.ids.spMonitor.values = monitores_cadastrados

        except FileNotFoundError:
            mostrar_popup('Arquivo não encontrado!', 'Arquivo de Cadastro de Monitores \nnão encontrado. Cadastre os \ndados faltantes nas suas sessões \nprimeiro.')
            return

    def ordenar_turmas_dia(self):
        self.ids.grid.clear_widgets()
        ordem_turmas = {}
        try:
            arq_slvOrdem = openpyxl.load_workbook('cadastro_OrdemTurmas.xlsx')
            celulas = arq_slvOrdem.active
            
            num = date.today().weekday()
            dia_semana_index = num + 1
            
            for celula in celulas.iter_rows(min_row=2):
                if dia_semana_index < len(celula):
                    ordem_turmas[celula[0].value] = celula[dia_semana_index].value

        except FileNotFoundError:
            mostrar_popup('Arquivo não encontrado!', 'Arquivo de Cadastro de Ordem \nnão encontrado. Cadastre os \ndados faltantes nas suas sessões \nprimeiro.')
            return

        ordem_turmas = {turma: ordem for turma, ordem in ordem_turmas.items() if ordem is not None}
        turmas_ordenadas = sorted(ordem_turmas.items(), key=lambda x: x[1])

        for ordem, (turma,_) in enumerate(turmas_ordenadas, start=1):
            form_turma = f'{ordem}º - {turma}'
            self.ids.grid.add_widget(Label(text=form_turma, size_hint_y=None, height=40))
            self.ids.grid.add_widget(TextInput(hint_text='Meninos', size_hint_y=None, height=40, multiline=False))
            self.ids.grid.add_widget(TextInput(hint_text='Meninas', size_hint_y=None, height=40, multiline=False))

    def salvar_frequencia(self, data, almoco, monitor, dia_semana):
        try:
            arq_turma = openpyxl.load_workbook('cadastro_turmas.xlsx')
            arq_ordemTurmas = openpyxl.load_workbook('cadastro_OrdemTurmas.xlsx')
            celulas_turma = arq_turma.active

            turmas_cadastradas = {}
            for celula in celulas_turma.iter_rows(min_row=2, max_col=2):
                turmas_cadastradas[celula[0].value] = celula[1].value

            if arq_ordemTurmas.active.max_row == 2:
                mostrar_popup('Erro ao Salvar!', 'Não é possível salvar a Frequência. \nArquivo de Cadastro de Ordem está \nvazio. Cadastre as turmas e suas \nrespectivas ordens primeiro.')
                return
        
        except FileNotFoundError:
            mostrar_popup('Erro ao Salvar!', 'Não é possível salvar a Frequência. \nArquivo de Cadastro de Ordem não \nencontrado. Cadastre as turmas \ne suas respectivas ordens primeiro.')
            return

        try:
            arq_slvFrequencia = openpyxl.load_workbook('frequencia.xlsx')
            celula_freq = arq_slvFrequencia.active

        except FileNotFoundError:
            arq_slvFrequencia = openpyxl.Workbook()
            celula_freq = arq_slvFrequencia.active
            celula_freq.append(['Data', 'Almoço', 'Monitor', 'Dia da Semana', 'Turmas', 'Meninos', 'Meninas'])

        divisao_grids = self.ids.grid.children
        num_turmas = len(divisao_grids) // 3

        for trm in range(num_turmas):
            turma_index = trm * 3 + 2
            turma = divisao_grids[turma_index].text[5:]

            if len(almoco) < 6 or len(almoco) > 50:
                mostrar_popup('Campo Vazio ou Valor Inválido!', 'Você digitou um almoço inválido \nou deixou o espaço em branco. \nDigite como informado no exemplo.')
                return
            
            if monitor == 'Escolha Monitor':
                mostrar_popup('Erro!', 'Nenhum Monitor foi escolhido. \nSelecione um Monitor.')
                return

            quantidade_meninos = divisao_grids[turma_index - 2].text
            quantidade_meninas = divisao_grids[turma_index - 1].text

            if not quantidade_meninos.isdigit() or not quantidade_meninas.isdigit():
                mostrar_popup('Campo Vazio!', 'Quantidade de meninos ou meninas \nnão está preenchida. Verifique todos \nos campos antes de salvar.')
                return

            quantidade_meninos = int(quantidade_meninos)
            quantidade_meninas = int(quantidade_meninas)
            total_alunos = quantidade_meninos + quantidade_meninas

            if turma in turmas_cadastradas and total_alunos > turmas_cadastradas[turma]:
                mostrar_popup('Valor Inválido!', f'A quantidade total de alunos na \nturma {turma} excede o limite\n cadastrado.')
                return
            
            celula_freq.append([data, almoco, monitor, dia_semana, turma, quantidade_meninos, quantidade_meninas])
        
        arq_slvFrequencia.save('frequencia.xlsx')
        self.ids.almoco.text = ''
        self.ids.spMonitor.text = 'Escolha Monitor'
        for i in range(len(divisao_grids)):
            if isinstance(divisao_grids[i], TextInput):
                divisao_grids[i].text = ''

        mostrar_popup('Sucesso!', 'Frequência salva com sucesso.')

class Relatorio(Screen):
    def on_pre_enter(self):
        self.ids.graficoSexo.clear_widgets()
        self.ids.tabMinMax.clear_widgets()
        self.ids.tabRanking.clear_widgets()

        self.carregar_dados()

    def on_enter(self):
        self.carregar_dados()
    
    def carregar_dados(self):
        try:
            arq_freq = openpyxl.load_workbook('frequencia.xlsx')
            celula_freq = arq_freq.active

        except FileNotFoundError:
            return mostrar_popup('Arquivo não encontrado!', 'Arquivo de frequência não encontrado. \nCadastre a frequência primeiro.')
        
        try:
            self.tab_min_max(celula_freq)
            self.tab_ranking(celula_freq)
            self.grafico_sexo(celula_freq)
        except Exception as e:
            mostrar_popup('Erro', f'Erro do tipo \n{e} \nao carregar os dados.')
            print(e)

    def grafico_sexo(self, celula_freq):
        def calcular_totais(inicio_semana, fim_semana):
            total_meninas = 0
            total_meninos = 0
            turmas_contabilizadas = {dia: set() for dia in ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta']}
            for col in celula_freq.iter_rows(min_row=2, values_only=True):
                data = col[0]
                dia_semana = col[3]
                turma = col[4]
                meninos = col[5]
                meninas = col[6]
            
                if isinstance(data, datetime):
                    data = data.date()
                elif isinstance(data, str):
                    data = datetime.strptime(data, '%d/%m/%Y').date()

                if inicio_semana <= data <= fim_semana and turma not in turmas_contabilizadas[dia_semana]:
                    turmas_contabilizadas[dia_semana].add(turma)

                    if isinstance(meninos, int):
                        total_meninos += meninos
                    if isinstance(meninas, int):
                        total_meninas += meninas
            return total_meninos, total_meninas

        hoje = datetime.now().date()
        inicio_semana = hoje - timedelta(days=hoje.weekday())
        fim_semana = inicio_semana + timedelta(days=4)
        total_meninos, total_meninas = calcular_totais(inicio_semana, fim_semana)

        if total_meninos == 0 and total_meninas == 0:
            inicio_semana = inicio_semana - timedelta(days=7)
            fim_semana = fim_semana - timedelta(days=7)
            total_meninos, total_meninas = calcular_totais(inicio_semana, fim_semana)

        sexo = ['Meninos', 'Meninas']
        dados = [total_meninos/5, total_meninas/5]
        cor = ['#7B68EE','#DDA0DD']

        def porcentagem(pct, allvals):
            absolute = int(pct/100.*np.sum(allvals))
            return '{:.1f}% \n({:d})'.format(pct, absolute)

        fig, ax = plt.subplots(figsize=(4, 1.5))
        ax.set_title('MÉDIA DA QUANTIDADE QUE \nALMOÇOU POR SEXO BIOLÓGICO', fontsize='8', color='#ffffff', fontweight='bold')
        fig.patch.set_facecolor('#B36699')

        wedges, texts, autotexts = ax.pie(dados, autopct=lambda pct: porcentagem(pct, dados), colors=cor)
        ax.legend(wedges, sexo, title='Sexo', loc='center left', bbox_to_anchor=(-0.15, 0, 0, 0), prop = {'size': 8})
        plt.setp(autotexts, size = 8)
        ax.axis('equal')

        self.ids.graficoSexo.clear_widgets()
        canvas = FigureCanvasKivyAgg(fig)
        canvas.size = (300, 300)
        canvas.pos_hint = {'center_x': 0.5, 'center_y': 0.5}
        self.ids.graficoSexo.add_widget(canvas)
    
    def tab_min_max(self, celula_freq):
        dias_semana = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta']
        qtde_pessoas = {dia: 0 for dia in dias_semana}
        turmas_contabilizadas = {dia: set() for dia in dias_semana}

        hoje = datetime.now().date()
        inicio_semana = hoje - timedelta(days=hoje.weekday()) 
        fim_semana = inicio_semana + timedelta(days=4)

        def calcular_qtdePessoas(inicio_semana, fim_semana):
            for row in celula_freq.iter_rows(min_row=2, values_only=True):
                data = row[0]
                dia_semana = row[3]
                turma = row[4]
                meninos = row[5]
                meninas = row[6]

                if isinstance(data, datetime):
                    data = data.date()
                elif isinstance(data, str):
                    data = datetime.strptime(data, '%d/%m/%Y').date()


                if inicio_semana <= data <= fim_semana and turma not in turmas_contabilizadas[dia_semana]:
                    turmas_contabilizadas[dia_semana].add(turma)
                    if dia_semana in qtde_pessoas:
                        qtde_pessoas[dia_semana] += meninos + meninas
        
        calcular_qtdePessoas(inicio_semana, fim_semana)

        if all(qtde == 0 for qtde in qtde_pessoas.values()):
            inicio_semana = inicio_semana - timedelta(days=7)
            fim_semana = fim_semana - timedelta(days=7)
            qtde_pessoas = {dia: 0 for dia in dias_semana}
            turmas_contabilizadas = {dia: set() for dia in dias_semana}
            calcular_qtdePessoas(inicio_semana, fim_semana)

        max_dia = max(qtde_pessoas, key=qtde_pessoas.get)
        min_dia = min(qtde_pessoas, key=qtde_pessoas.get)

        dados = [
            ('Dia', 'Qtde'),
            (max_dia, qtde_pessoas[max_dia]), 
            (min_dia, qtde_pessoas[min_dia])
        ]
        fig, ax = plt.subplots(figsize=(3, 1.5))
        ax.set_title('QUANTIDADE MÍNIMA E \nMÁXIMA DA SEMANA', fontsize=8, color='#ffffff', fontweight='bold', pad=1)
        ax.axis('tight')
        ax.axis('off')
        fig.patch.set_facecolor('#B36699')

        table = ax.table(cellText=list(dados[1:]), colLabels=dados[0], cellLoc='center', loc='center')
        table.auto_set_font_size(False)
        table.set_fontsize(8.5)

        self.ids.tabMinMax.clear_widgets()
        canvas = FigureCanvasKivyAgg(fig)
        canvas.size = (300, 300)
        canvas.pos_hint = {'center_x': 0.5, 'center_y': 0.5}
        self.ids.tabMinMax.add_widget(canvas)

    def tab_ranking(self, celula_freq):
        try:
            arq_turma = openpyxl.load_workbook('cadastro_turmas.xlsx')
            celulas_turma = arq_turma.active
        except FileNotFoundError:
            return

        turmas_totais = {}
        for celula in celulas_turma.iter_rows(min_row=2, max_col=2):
            turmas_totais[celula[0].value] = celula[1].value

        turmas_alunos = {}
        for row in celula_freq.iter_rows(min_row=2, values_only=True):
            turma = row[4]
            meninos = row[5]
            meninas = row[6]
            total_alunos = meninos + meninas
            if turma in turmas_alunos:
                turmas_alunos[turma] = max(turmas_alunos[turma], total_alunos)
            else:
                turmas_alunos[turma] = total_alunos

        turmas_porcentagem = {}
        for turma, alunos in turmas_alunos.items():
            if turma in turmas_totais:
                turmas_porcentagem[turma] = (alunos / turmas_totais[turma]) * 100

        turmas_ordenadas = sorted(turmas_porcentagem.items(), key=lambda x: x[1], reverse=True)[:3]

        dados = [('Turma', '% Alunos')] + [(turma, f'{porcentagem:.2f}%') for turma, porcentagem in turmas_ordenadas]
        fig, ax = plt.subplots(figsize=(3, 1.5))
        ax.set_title(' RANKING DAS TURMAS \nQUE MAIS ALMOÇARAM', fontsize=8, color='#ffffff', fontweight='bold', pad=1)
        ax.axis('tight')
        ax.axis('off')
        fig.patch.set_facecolor('#B36699')

        table = ax.table(cellText=list(dados[1:]), colLabels=dados[0], cellLoc='center', loc='center')
        table.auto_set_font_size(False)
        table.set_fontsize(8.5)

        self.ids.tabRanking.clear_widgets()
        canvas = FigureCanvasKivyAgg(fig)
        canvas.size = (300, 300) 
        canvas.pos_hint = {'center_x': 0.5, 'center_y': 0.5}
        self.ids.tabRanking.add_widget(canvas)

class Mofome(App):
    def build(self):
        Window.size = (369, 640)
        Window.set_icon('logo.png')
        return Gerenciador()

if __name__ == '__main__':
    Mofome().run()
